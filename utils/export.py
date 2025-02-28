import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database.models import Task, User, Implementation, Offer

def export_tasks_to_excel(tasks, file_path):
    """
    Eksportuje listę zadań do pliku Excel
    
    Args:
        tasks (list): Lista zadań do eksportu
        file_path (str): Ścieżka do pliku wynikowego
        
    Returns:
        bool: True jeśli eksport się powiódł, False w przeciwnym przypadku
    """
    try:
        # Utwórz nowy workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Zadania"
        
        # Ustaw nagłówki
        headers = ["ID", "Użytkownik", "Kategoria", "Typ", "Opis", 
                   "Czas rozpoczęcia", "Czas zakończenia", "Czas trwania (min)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            # Formatowanie nagłówka
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Wypełnij danymi
        for row, task in enumerate(tasks, 2):
            # Pobierz dane użytkownika
            user = User.get_by_id(task.user_id)
            user_name = f"{user.first_name} {user.last_name}" if user else "Nieznany"
            
            # Dodaj wiersz
            ws.cell(row=row, column=1).value = task.id
            ws.cell(row=row, column=2).value = user_name
            ws.cell(row=row, column=3).value = task.category
            ws.cell(row=row, column=4).value = task.task_type
            ws.cell(row=row, column=5).value = task.description
            ws.cell(row=row, column=6).value = task.start_time
            ws.cell(row=row, column=7).value = task.end_time
            ws.cell(row=row, column=8).value = task.duration
            
            # Dodaj dodatkowe informacje dla zadań typu Wdrożenie lub Oferta
            if task.task_type == "Wdrożenie" and task.implementation_id:
                impl = Implementation.get_by_id(task.implementation_id)
                if impl:
                    ws.cell(row=row, column=5).value = f"{task.description} (Wdrożenie: {impl.name})"
            
            elif task.task_type == "Oferta" and task.offer_id:
                offer = Offer.get_by_id(task.offer_id)
                if offer:
                    ws.cell(row=row, column=5).value = f"{task.description} (Oferta: {offer.name})"
        
        # Formatowanie tabeli
        for row in range(2, len(tasks) + 2):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Dostosuj szerokość kolumn
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Kolumna opisu szersza
        ws.column_dimensions['E'].width = 30
        
        # Automatyczne filtrowanie
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(tasks) + 1}"
        
        # Upewnij się, że katalog docelowy istnieje
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        
        # Zapisz plik
        wb.save(file_path)
        return True
    
    except Exception as e:
        print(f"Błąd podczas eksportu do Excel: {e}")
        return False

def export_implementations_to_excel(implementations, file_path, append=False):
    """
    Eksportuje listę wdrożeń do pliku Excel
    
    Args:
        implementations (list): Lista wdrożeń do eksportu
        file_path (str): Ścieżka do pliku wynikowego
        append (bool): Czy dołączyć do istniejącego pliku
        
    Returns:
        bool: True jeśli eksport się powiódł, False w przeciwnym przypadku
    """
    try:
        # Utwórz nowy workbook lub otwórz istniejący
        if append:
            try:
                wb = openpyxl.load_workbook(file_path)
                # Sprawdź czy arkusz "Wdrożenia" już istnieje, jeśli tak, usuń go
                if "Wdrożenia" in wb.sheetnames:
                    del wb["Wdrożenia"]
            except:
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
        else:
            wb = openpyxl.Workbook()
        
        ws = wb.create_sheet("Wdrożenia")
        
        # Ustaw nagłówki
        headers = ["ID", "Nazwa", "Opis", "Status", "Data rozpoczęcia", "Data zakończenia", 
                   "Wdrożenie (Użytkownik)", "Spawanie (Użytkownik)", "Malowanie (Użytkownik)", 
                   "Klejenie (Użytkownik)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            # Formatowanie nagłówka
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Wypełnij danymi
        for row, impl in enumerate(implementations, 2):
            # Pobierz daty głównej operacji
            start_date = ""
            end_date = ""
            if "Wdrożenie" in impl.operations:
                start_date = impl.operations["Wdrożenie"].get("start_date", "")
                end_date = impl.operations["Wdrożenie"].get("end_date", "")
            
            # Podstawowe informacje
            ws.cell(row=row, column=1).value = impl.id
            ws.cell(row=row, column=2).value = impl.name
            ws.cell(row=row, column=3).value = impl.description
            ws.cell(row=row, column=4).value = impl.status
            ws.cell(row=row, column=5).value = start_date
            ws.cell(row=row, column=6).value = end_date
            
            # Operacje - tylko informacje o użytkownikach
            col_offset = 7
            for i, operation in enumerate(["Wdrożenie", "Spawanie", "Malowanie", "Klejenie"]):
                op_data = impl.operations.get(operation, {})
                user_id = op_data.get('user_id')
                user = User.get_by_id(user_id) if user_id else None
                user_name = f"{user.first_name} {user.last_name}" if user else ""
                
                ws.cell(row=row, column=col_offset + i).value = user_name
        
        # Formatowanie tabeli
        for row in range(2, len(implementations) + 2):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Dostosuj szerokość kolumn
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Kolumna opisu szersza
        ws.column_dimensions['C'].width = 30
        
        # Upewnij się, że katalog docelowy istnieje
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        
        # Zapisz plik
        wb.save(file_path)
        return True
    
    except Exception as e:
        print(f"Błąd podczas eksportu do Excel: {e}")
        return False

def export_offers_to_excel(offers, file_path, append=False):
    """
    Eksportuje listę ofert do pliku Excel
    
    Args:
        offers (list): Lista ofert do eksportu
        file_path (str): Ścieżka do pliku wynikowego
        append (bool): Czy dołączyć do istniejącego pliku
        
    Returns:
        bool: True jeśli eksport się powiódł, False w przeciwnym przypadku
    """
    try:
        # Utwórz nowy workbook lub otwórz istniejący
        if append:
            try:
                wb = openpyxl.load_workbook(file_path)
                # Sprawdź czy arkusz "Oferty" już istnieje, jeśli tak, usuń go
                if "Oferty" in wb.sheetnames:
                    del wb["Oferty"]
            except:
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
        else:
            wb = openpyxl.Workbook()
            
        ws = wb.create_sheet("Oferty")
        
        # Ustaw nagłówki
        headers = ["ID", "Nazwa", "Opis", "Status", "Data rozpoczęcia", "Data zakończenia", 
                   "Wdrożenie (Użytkownik)", "Spawanie (Użytkownik)", "Malowanie (Użytkownik)", 
                   "Klejenie (Użytkownik)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            # Formatowanie nagłówka
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Wypełnij danymi
        for row, offer in enumerate(offers, 2):
            # Pobierz daty głównej operacji
            start_date = ""
            end_date = ""
            if "Wdrożenie" in offer.operations:
                start_date = offer.operations["Wdrożenie"].get("start_date", "")
                end_date = offer.operations["Wdrożenie"].get("end_date", "")
            
            # Podstawowe informacje
            ws.cell(row=row, column=1).value = offer.id
            ws.cell(row=row, column=2).value = offer.name
            ws.cell(row=row, column=3).value = offer.description
            ws.cell(row=row, column=4).value = offer.status
            ws.cell(row=row, column=5).value = start_date
            ws.cell(row=row, column=6).value = end_date
            
            # Operacje - tylko informacje o użytkownikach
            col_offset = 7
            for i, operation in enumerate(["Wdrożenie", "Spawanie", "Malowanie", "Klejenie"]):
                op_data = offer.operations.get(operation, {})
                user_id = op_data.get('user_id')
                user = User.get_by_id(user_id) if user_id else None
                user_name = f"{user.first_name} {user.last_name}" if user else ""
                
                ws.cell(row=row, column=col_offset + i).value = user_name
        
        # Formatowanie tabeli
        for row in range(2, len(offers) + 2):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Dostosuj szerokość kolumn
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Kolumna opisu szersza
        ws.column_dimensions['C'].width = 30
        
        # Upewnij się, że katalog docelowy istnieje
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        
        # Zapisz plik
        wb.save(file_path)
        return True
    
    except Exception as e:
        print(f"Błąd podczas eksportu do Excel: {e}")
        return False